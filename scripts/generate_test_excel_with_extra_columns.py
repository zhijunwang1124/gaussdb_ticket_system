# -*- coding: utf-8 -*-
"""
生成包含完整字段列表和额外测试列的100条样例数据Excel文件
用于测试导入时是否能正确忽略不匹配的列
"""
import random
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("请先安装: pip install openpyxl")
    exit(1)

# 完整的业务字段列表（46个）
TICKET_COLUMNS = [
    "当前阶段",
    "问题严重性",
    "局点",
    "实例简称",
    "业务名称",
    "当前处理人",
    "描述",
    "滞留时间",
    "SLA时间",
    "问题审核人",
    "运维人员分析人",
    "开发人员分析人",
    "开发人员闭环人",
    "运维人员闭环人",
    "问题审核关闭人",
    "进展概述",
    "创建日期",
    "问题审核总时长",
    "运维人员分析总时长",
    "开发人员分析总时长",
    "开发人员闭环总时长",
    "运维人员闭环总时长",
    "问题审核关闭总时长",
    "管控版本",
    "对外答复",
    "故障到恢复用时",
    "是否涉及故障恢复",
    "磐石版本是否涉及",
    "是否需要预警",
    "问题根因",
    "DTS单号",
    "规避措施",
    "是否质量问题",
    "恢复方法",
    "是否咨询问题",
    "是否涉及内核升级",
    "协同处理人",
    "高斯版本",
    "HCS版本号",
    "HCS/轻量化",
    "问题类型",
    "根因分类",
    "部署形态"
]

# 额外的测试列（应该被忽略）
EXTRA_COLUMNS = [
    "备注信息",
    "内部编号",
    "处理优先级",
    "业务部门",
    "客户联系人",
    "预计处理时间",
    "实际处理时间",
    "满意度评分",
    "是否需要升级",
    "临时标记",
    "旧版本工单号",
    "备用联系方式"
]

# 测试数据
PHASES = ["问题审核", "运维人员分析", "开发人员分析", "开发人员闭环", "运维人员闭环", "问题关闭"]
SEVERITIES = ["严重", "较严重", "一般", "轻微"]
SITES = ["北京局", "上海局", "深圳局", "杭州局", "成都局", "广州局"]
INSTANCE_NAMES = ["inst001", "inst002", "inst003", "inst004", "inst005", "inst006"]
BUSINESS_NAMES = ["核心业务", "交易系统", "报表系统", "网关服务", "监控系统"]
HANDLERS = ["张三", "李四", "王五", "赵六", "钱七", "孙八"]
DURATIONS = ["2小时", "4小时", "8小时", "16小时", "24小时", "48小时", "72小时"]
CTRL_VERSIONS = ["V2.1.0", "V2.2.0", "V2.3.0", "V3.0.0", "V3.1.0", "V3.2.0"]
KERNEL_VERSIONS = ["K505.1", "K506.0", "K507.1", "K508.0", "K509.0", "K510.1"]
HCS_VERSIONS = ["HCS 8.0.1", "HCS 8.0.2", "HCS 8.0.3", "HCS 8.1.0"]
HCS_TYPES = ["HCS", "轻量化"]
PROBLEM_TYPES = ["性能问题", "功能缺陷", "配置问题", "网络问题", "资源问题"]
ROOT_CAUSE_CATEGORIES = ["代码缺陷", "配置错误", "资源不足", "网络故障", "第三方问题"]
DEPLOY_TYPES = ["集中式", "分布式"]
YES_NO = ["是", "否"]
RECOVERY_METHODS = ["自动恢复", "人工干预", "服务重启", "回滚版本"]
WORKAROUNDS = ["临时规避方案", "重启服务", "参数调整", "扩容资源"]

PRIORITIES = ["高", "中", "低"]
DEPARTMENTS = ["研发部", "运维部", "产品部", "测试部"]

def generate_yw_no(index):
    """生成工单号"""
    return f"YW{8800000000 + index}"

def random_date(start_year=2025, months_range=12):
    """生成随机日期"""
    start = datetime(start_year, 1, 1)
    end = start + timedelta(days=months_range * 30)
    delta = end - start
    random_days = random.randint(0, delta.days)
    return start + timedelta(days=random_days)

def generate_sample_excel(filename="ticket_test_sample_with_extra_columns.xlsx"):
    """生成包含完整字段和额外列的测试Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工单数据"

    # 写入表头
    headers = ["工单号"] + TICKET_COLUMNS + EXTRA_COLUMNS
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 生成100条测试数据
    random.seed(42)
    for i in range(1, 101):
        row = i + 1
        yw_no = generate_yw_no(i - 1)
        create_date = random_date(2025, 12)
        date_str = create_date.strftime("%Y-%m-%d")

        # 基础数据
        phase = random.choice(PHASES)
        severity = random.choice(SEVERITIES)
        site = random.choice(SITES)
        instance_name = random.choice(INSTANCE_NAMES)
        business_name = random.choice(BUSINESS_NAMES)
        handler = random.choice(HANDLERS)
        ctrl_ver = random.choice(CTRL_VERSIONS)
        kernel_ver = random.choice(KERNEL_VERSIONS)
        hcs_ver = random.choice(HCS_VERSIONS)
        hcs_type = random.choice(HCS_TYPES)
        problem_type = random.choice(PROBLEM_TYPES)
        root_cause_category = random.choice(ROOT_CAUSE_CATEGORIES)
        deploy_type = random.choice(DEPLOY_TYPES)

        # 构建行数据
        data = [
            yw_no,  # 工单号
            phase,  # 当前阶段
            severity,  # 问题严重性
            site,  # 局点
            instance_name,  # 实例简称
            business_name,  # 业务名称
            handler,  # 当前处理人
            f"问题描述样本{i}：GaussDB 连接异常或查询超时相关描述",  # 描述
            random.choice(DURATIONS),  # 滞留时间
            random.choice(DURATIONS),  # SLA时间
            handler,  # 问题审核人
            handler,  # 运维人员分析人
            handler,  # 开发人员分析人
            handler,  # 开发人员闭环人
            handler,  # 运维人员闭环人
            handler,  # 问题审核关闭人
            f"进展概述{i}：已联系现场排查网络与参数",  # 进展概述
            date_str,  # 创建日期
            random.choice(DURATIONS),  # 问题审核总时长
            random.choice(DURATIONS),  # 运维人员分析总时长
            random.choice(DURATIONS),  # 开发人员分析总时长
            random.choice(DURATIONS),  # 开发人员闭环总时长
            random.choice(DURATIONS),  # 运维人员闭环总时长
            random.choice(DURATIONS),  # 问题审核关闭总时长
            ctrl_ver,  # 管控版本
            f"对外答复{i}：已处理并建议后续观察",  # 对外答复
            random.choice(DURATIONS),  # 故障到恢复用时
            random.choice(YES_NO),  # 是否涉及故障恢复
            random.choice(YES_NO),  # 磐石版本是否涉及
            random.choice(YES_NO),  # 是否需要预警
            f"问题根因{i}：与配置/资源/网络相关的示例根因",  # 问题根因
            f"DTS{i:06d}",  # DTS单号
            f"规避措施{i}：临时规避方案",  # 规避措施
            random.choice(YES_NO),  # 是否质量问题
            random.choice(RECOVERY_METHODS),  # 恢复方法
            random.choice(YES_NO),  # 是否咨询问题
            random.choice(YES_NO),  # 是否涉及内核升级
            "张三,李四",  # 协同处理人
            kernel_ver,  # 高斯版本
            hcs_ver,  # HCS版本号
            hcs_type,  # HCS/轻量化
            problem_type,  # 问题类型
            root_cause_category,  # 根因分类
            deploy_type,  # 部署形态
        ]

        # 添加额外的测试列数据
        data.extend([
            f"备注信息{i}：此为测试备注",  # 备注信息
            f"INTERNAL-{i:08d}",  # 内部编号
            random.choice(PRIORITIES),  # 处理优先级
            random.choice(DEPARTMENTS),  # 业务部门
            f"客户{i}",  # 客户联系人
            f"{random.randint(1, 24)}小时",  # 预计处理时间
            f"{random.randint(1, 24)}小时",  # 实际处理时间
            random.randint(1, 5),  # 满意度评分
            random.choice(YES_NO),  # 是否需要升级
            f"TEMP-{i:04d}",  # 临时标记
            f"OLD-{20250000 + i}",  # 旧版本工单号
            f"13800138{i:03d}",  # 备用联系方式
        ])

        # 写入行数据
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row, column=col_idx, value=value)

    # 保存文件
    wb.save(filename)
    print(f"[OK] 已生成测试文件: {filename}")
    print(f"[OK] 共生成100条测试数据")
    print(f"[OK] 包含 {len(TICKET_COLUMNS)} 个标准业务字段")
    print(f"[OK] 包含 {len(EXTRA_COLUMNS)} 个额外测试列（应该被忽略）")
    print(f"[OK] 总列数: {len(headers)}")
    print("\n字段列表：")
    print(f"标准字段 ({len(TICKET_COLUMNS)} 个):")
    for idx, col in enumerate(TICKET_COLUMNS, 1):
        print(f"  {idx:2d}. {col}")
    print(f"\n额外列 ({len(EXTRA_COLUMNS)} 个，应该被忽略):")
    for idx, col in enumerate(EXTRA_COLUMNS, 1):
        print(f"  {idx:2d}. {col}")

if __name__ == "__main__":
    generate_sample_excel()
    print("\n[OK] 测试文件生成完成！")
    print("[OK] 可以使用此文件测试导入功能，验证额外列是否被正确忽略")